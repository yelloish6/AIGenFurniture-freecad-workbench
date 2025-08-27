import openpyxl
import shutil, os
from openpyxl.styles.borders import Border, Side


def export_pal_for_proficut(order, output_folder):

    file_path = output_folder + "/Comanda_PAL_" + order.mat_pal + "_" + order.client + ".xlsx"
    template_path = os.path.join(os.path.dirname(__file__), "templates", "Cote-Proficut-2018.xlsx")
    shutil.copyfile(template_path, file_path)
    # file = openpyxl.load_workbook(file_path)
    file = openpyxl.load_workbook(file_path, data_only=True, keep_links=False)
    # clean up images if template has any
    file._images = []
    for ws in file.worksheets:
        ws._images = []
    sheet = file.get_sheet_by_name("Sheet1")

    sheet['C1'] = order.client_proficut
    sheet['D2'] = order.tel_proficut
    sheet['D3'] = order.transport
    sheet['C4'] = order.address
    sheet['G4'] = order.mat_pal

    counter = 0
    for corp in order.cabinets_list:
        for element in corp.elements_list:
            if element.type == "pal":
                sheet['A' + str(10 + counter)] = "1"
                sheet['B' + str(10 + counter)] = element.length
                sheet['C' + str(10 + counter)] = element.width
                sheet['D' + str(10 + counter)] = 0
                sheet['E' + str(10 + counter)] = element.label
                if element.cant_list[0] == 0.4:
                    sheet['F' + str(10 + counter)] = 1
                else:
                    sheet['F' + str(10 + counter)] = element.cant_list[0]
                if element.cant_list[1] == 0.4:
                    sheet['G' + str(10 + counter)] = 1
                else:
                    sheet['G' + str(10 + counter)] = element.cant_list[1]
                if element.cant_list[2] == 0.4:
                    sheet['H' + str(10 + counter)] = 1
                else:
                    sheet['H' + str(10 + counter)] = element.cant_list[2]
                if element.cant_list[3] == 0.4:
                    sheet['I' + str(10 + counter)] = 1
                else:
                    sheet['I' + str(10 + counter)] = element.cant_list[3]
                sheet['K' + str(10 + counter)] = element.obs
                counter += 1

    # draw the sketch template for cut-out boards (L-Shape, U-Shape)
        # L-Shape - left
        sheet_2 = file.get_sheet_by_name("Sheet2")

        sheet_2['B2'] = 'label'

        sheet_2['B4'].border = Border(top=Side(style='thick'), left=Side(style='thick'))
        sheet_2['C4'].border = Border(top=Side(style='thick'))
        sheet_2['D4'].border = Border(top=Side(style='thick'))
        sheet_2['E4'].border = Border(top=Side(style='thick'))
        sheet_2['F4'].border = Border(top=Side(style='thick'), right=Side(style='thick'))

        sheet_2['B5'].border = Border(left=Side(style='thick'))
        sheet_2['B6'].border = Border(left=Side(style='thick'))
        sheet_2['B7'].border = Border(left=Side(style='thick'))
        sheet_2['B8'].border = Border(left=Side(style='thick'))
        sheet_2['B9'].border = Border(left=Side(style='thick'), bottom=Side(style='thick'))

        sheet_2['C9'].border = Border(bottom=Side(style='thick'), right=Side(style='thick'))
        sheet_2['C8'].border = Border(right=Side(style='thick'))
        sheet_2['C7'].border = Border(right=Side(style='thick'))

        sheet_2['D6'].border = Border(bottom=Side(style='thick'))
        sheet_2['E6'].border = Border(bottom=Side(style='thick'))
        sheet_2['F6'].border = Border(bottom=Side(style='thick'), right=Side(style='thick'))

        sheet_2['F5'].border = Border(right=Side(style='thick'))

        sheet_2['D3'] = "cota 1"
        sheet_2['G5'] = "cota 2"
        sheet_2['E7'] = "cota 3"
        sheet_2['D8'] = "cota 4"
        sheet_2['B10'] = "cota 5"
        sheet_2['A7'] = "cota 6"

        # L-Shape - right

        sheet_2['B12'] = 'label'

        sheet_2['B14'].border = Border(top=Side(style='thick'), left=Side(style='thick'))
        sheet_2['C14'].border = Border(top=Side(style='thick'))
        sheet_2['D14'].border = Border(top=Side(style='thick'))
        sheet_2['E14'].border = Border(top=Side(style='thick'))
        sheet_2['F14'].border = Border(top=Side(style='thick'), right=Side(style='thick'))

        sheet_2['B15'].border = Border(left=Side(style='thick'))
        sheet_2['B16'].border = Border(left=Side(style='thick'), bottom=Side(style='thick'))
        sheet_2['C16'].border = Border(bottom=Side(style='thick'))
        sheet_2['D16'].border = Border(bottom=Side(style='thick'))

        sheet_2['E17'].border = Border(left=Side(style='thick'))
        sheet_2['E18'].border = Border(left=Side(style='thick'))

        sheet_2['E19'].border = Border(bottom=Side(style='thick'), left=Side(style='thick'))
        sheet_2['F19'].border = Border(bottom=Side(style='thick'), right=Side(style='thick'))

        sheet_2['F15'].border = Border(right=Side(style='thick'))
        sheet_2['F16'].border = Border(right=Side(style='thick'))
        sheet_2['F17'].border = Border(right=Side(style='thick'))
        sheet_2['F18'].border = Border(right=Side(style='thick'))

        sheet_2['D13'] = "cota 1"
        sheet_2['G16'] = "cota 2"
        sheet_2['E20'] = "cota 3"
        sheet_2['D18'] = "cota 4"
        sheet_2['C17'] = "cota 5"
        sheet_2['A15'] = "cota 6"

        # U-Shape

        sheet_2['B22'] = 'label'

        sheet_2['B24'].border = Border(top=Side(style='thick'), left=Side(style='thick'), right=Side(style='thick'))
        sheet_2['B25'].border = Border(left=Side(style='thick'), right=Side(style='thick'))
        sheet_2['B26'].border = Border(left=Side(style='thick'), right=Side(style='thick'))
        sheet_2['B27'].border = Border(left=Side(style='thick'), right=Side(style='thick'))
        sheet_2['B28'].border = Border(left=Side(style='thick'))
        sheet_2['B29'].border = Border(left=Side(style='thick'), bottom=Side(style='thick'))

        sheet_2['F24'].border = Border(top=Side(style='thick'), left=Side(style='thick'), right=Side(style='thick'))
        sheet_2['F25'].border = Border(left=Side(style='thick'), right=Side(style='thick'))
        sheet_2['F26'].border = Border(left=Side(style='thick'), right=Side(style='thick'))
        sheet_2['F27'].border = Border(left=Side(style='thick'), right=Side(style='thick'))
        sheet_2['F28'].border = Border(right=Side(style='thick'))
        sheet_2['F29'].border = Border(right=Side(style='thick'), bottom=Side(style='thick'))

        sheet_2['C27'].border = Border(bottom=Side(style='thick'))
        sheet_2['D27'].border = Border(bottom=Side(style='thick'))
        sheet_2['E27'].border = Border(bottom=Side(style='thick'))

        sheet_2['C29'].border = Border(bottom=Side(style='thick'))
        sheet_2['D29'].border = Border(bottom=Side(style='thick'))
        sheet_2['E29'].border = Border(bottom=Side(style='thick'))

        sheet_2['B23'] = "cota 1"
        sheet_2['C25'] = "cota 2"
        sheet_2['D27'] = "cota 3"
        sheet_2['E25'] = "cota 4"
        sheet_2['F23'] = "cota 5"
        sheet_2['G26'] = "cota 6"
        sheet_2['D30'] = "cota 7"
        sheet_2['A26'] = "cota 8"

        file.save(file_path)


def export_pfl_for_proficut(order, output_folder):

    file_path = output_folder + "/Comanda_PFL_" + order.mat_pfl + "_" + order.client + ".xlsx"
    template_path = os.path.join(os.path.dirname(__file__), "templates", "Cote-Proficut-2018.xlsx")
    shutil.copyfile(template_path, file_path)

    file = openpyxl.load_workbook(file_path)
    sheet = file.get_sheet_by_name("Sheet1")

    sheet['C1'] = order.client_proficut
    sheet['D2'] = order.tel_proficut
    sheet['D3'] = order.transport
    sheet['C4'] = order.address
    sheet['G4'] = order.mat_pfl

    counter = 0
    for cabinet in order.cabinets_list:
        for element in cabinet.elements_list:
            if element.type == "pfl":
                sheet['A' + str(10 + counter)] = "1"
                sheet['B' + str(10 + counter)] = element.length
                sheet['C' + str(10 + counter)] = element.width
                sheet['D' + str(10 + counter)] = 0
                sheet['E' + str(10 + counter)] = element.label
                counter += 1
    file.save(file_path)
